"""
Конвертирует прайс-лист (public/prices/price-current.xlsx) в структурированные данные
для интерактивного списка на сайте (src/data/priceItems.json).

Запуск при каждом обновлении прайса: python scripts/xlsx-to-price-items.py
Формат исходника: колонка B — название бренда/линейки (строка-заголовок) или товара,
колонка C — цена, колонка D — единица измерения.

Заголовки — два уровня, различаются заливкой ячейки (индексированный цвет Excel):
  - indexed=8  (чёрный) — бренд (напр. WELLA, SCHWARZKOPF)
  - indexed=22 (серый)  — линейка внутри бренда (напр. OSIS, SILHOUETTE)
Товары, выделенные красным шрифтом в Excel, помечаются promo: true (для блока «Акции»).
"""
import json
import re
from datetime import date, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC_XLSX = ROOT / "public" / "prices" / "price-current.xlsx"
OUT_JSON = ROOT / "src" / "data" / "priceItems.json"
# Сайдкар из build_price_current.py (Price/ - в .gitignore, не деплоится): старая цена
# и скидка по каждому акционному товару - price-current.xlsx хранит только цену со
# скидкой, старая цена иначе теряется. Файла может не быть (напр. кто-то прогнал этот
# скрипт отдельно, без пересборки прайса) - тогда просто не добавляем oldPrice/discountPct.
PROMO_META_PATH = ROOT / "Price" / "promo-meta.json"
# Ведётся В РЕПОЗИТОРИИ (не в Price/, который локальный и в .gitignore) - в отличие от
# промо-сайдкара, это НАКАПЛИВАЕМЫЙ журнал "когда товар впервые встретился в прайсе",
# должен пережить любую пересборку и не потеряться со сменой машины/сессии. Формат:
# {имя_товара: "YYYY-MM-DD" первого появления}. НЕ удалять и не редактировать руками -
# один раз потерянный файл заставит все текущие товары ошибочно выглядеть "новыми".
FIRST_SEEN_PATH = ROOT / "src" / "data" / "price-first-seen.json"
NEW_WINDOW_DAYS = 14

# Категории — журнал "название товара -> номер категории" В РЕПОЗИТОРИИ (не в Price/),
# заведён 2026-09-03 по прямой просьбе пользователя: он вручную проставляет категории
# прямо в колонке F price-current.xlsx, но build_price_current.py регенерирует этот файл
# из свежих 1С-исходников и ничего не знает о ручной разметке - без этого журнала она бы
# терялась при каждом обновлении прайса. Источник истины при каждом запуске - сама
# колонка F текущего price-current.xlsx (пользователь мог поменять категорию), журнал
# только переживает пересборки. build_price_current.py читает этот же файл, чтобы
# заранее проставить уже известные категории в новый прайс и подсветить синим то, чего
# в журнале ещё нет (см. Price/build_price_current.py, add_category_column).
CATEGORIES_PATH = ROOT / "src" / "data" / "price-categories.json"
CATEGORY_NAMES = {
    1: "Аммиачные красители",
    2: "Безаммиачные красители",
    3: "Окислители и осветляющие средства",
    4: "Технические средства",
    5: "Химия волос",
    6: "Ампулы/концентраты",
    7: "Брови и ресницы",
    8: "Шампуни",
    9: "Кондиционеры",
    10: "Маски",
    11: "Несмываемый уход",
    12: "Термозащита",
    13: "Лаки",
    14: "Спреи",
    15: "Пенки",
    16: "Крема и гели для укладки",
    17: "Расходные материалы",
}

# Остатки на складе (Price/Остатки_*.xlsx, выгрузка 1С) — для честного "Лучшая скидка" в
# каталоге (сортировка по количеству на складе, а не только по проценту скидки). Полностью
# автоматически: сам находит самый свежий файл по дате в имени, отдельно запускать ничего
# не нужно. Если файлов ещё нет — просто не добавляем поле stock, ничего не ломается.
STOCK_DATE_RE = re.compile(r"Остатки_(\d{2}),(\d{2}),(\d{4})")

BRAND_FILL_INDEX = 8
LINE_FILL_INDEX = 22

# Служебные строки 1С (не товар) - никогда не должны попадать в каталог сайта.
# См. Site/CLAUDE.md / Site/Price/build_price_current.py (та же логика, зафиксировано 2026-08-25).
EXCLUDE_NAME_PREFIXES = ["доставка"]
EXCLUDE_NAME_SUBSTRINGS = ["мятая", "мятые"]  # брак/мятая упаковка
EXCLUDE_ARTICLES = {"007927"}
JUNK_SECTION_SUFFIXES = ("_акция",)  # секции вроде "Wella_Акция"/"Londa_Акция" - клиренс без цены


def article(name: str):
    toks = name.strip().split()
    return toks[-1] if toks else None


def is_excluded(name: str) -> bool:
    n = name.strip().lower()
    if any(n.startswith(p) for p in EXCLUDE_NAME_PREFIXES):
        return True
    if any(s in n for s in EXCLUDE_NAME_SUBSTRINGS):
        return True
    if article(name) in EXCLUDE_ARTICLES:
        return True
    return False


def is_junk_section_header(name: str) -> bool:
    n = name.strip().lower()
    return any(n.endswith(suf) for suf in JUNK_SECTION_SUFFIXES)


def is_red(cell) -> bool:
    try:
        rgb = cell.font.color.rgb
        return isinstance(rgb, str) and rgb.upper().endswith("FF0000")
    except Exception:
        return False


def header_level(cell) -> str:
    """Возвращает 'brand', 'line' или 'unknown' по заливке ячейки-заголовка."""
    fg = cell.fill.fgColor
    if fg.type == "indexed":
        if fg.indexed == BRAND_FILL_INDEX:
            return "brand"
        if fg.indexed == LINE_FILL_INDEX:
            return "line"
    return "unknown"


def load_stock_levels(items: list) -> dict:
    """{артикул: количество} из самого свежего Price/Остатки_*.xlsx — только для АРТИКУЛОВ,
    однозначных сразу в ДВУХ местах:
    1) в самом priceItems.json — этот артикул встречается только у ОДНОГО бренда (короткие
       коды оттенков вида "0/11" реально совпадают у разных товаров разных брендов);
    2) в самом файле остатков — этот артикул встречается только под ОДНИМ заголовком секции.

    Почему не просто "бренд+артикул": иерархия заголовков в файле остатков НЕ совпадает 1:1
    со структурой бренд/линейка прайса (реальный случай 2026-09-04 — линейка "Total Results
    New" у Matrix идёт в остатках отдельной секцией верхнего уровня, без родителя "MATRIX",
    из-за чего сопоставление по (бренд-из-остатков, артикул) давало только ~53% совпадений
    вместо реальных ~94%). Двойная проверка неоднозначности проще и безопаснее, чем пытаться
    воспроизвести чужую иерархию заголовков: если код используется больше одного раза в
    любом из двух источников — просто не трогаем его, не гадаем.
    """
    candidates = []
    for p in (ROOT / "Price").glob("Остатки_*.xlsx"):
        m = STOCK_DATE_RE.search(p.name)
        if not m:
            continue
        d = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        candidates.append((d, p.stat().st_mtime, p))
    if not candidates:
        print("ВНИМАНИЕ: Price/Остатки_*.xlsx не найден - товары останутся без поля stock (не влияет на цены/каталог).")
        return {}
    candidates.sort(key=lambda t: (t[0], t[1]))
    src = candidates[-1][2]

    brands_by_article: dict = {}
    for it in items:
        a = article(it["name"])
        brands_by_article.setdefault(a, set()).add(it["brand"])
    price_unambiguous = {a for a, brands in brands_by_article.items() if len(brands) == 1}

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Sheet1"]
    qty_by_article: dict = {}
    headers_by_article: dict = {}
    header = None
    for row in ws.iter_rows(min_row=9, values_only=True):
        if not row or len(row) < 6:
            continue
        name, _, _unit, _cost, qty, _total = row[:6]
        if not name:
            continue
        # Заголовок секции — у товарной строки количество всегда число, у заголовка это
        # пусто/строка-заполнитель (не полагаемся на колонку единиц измерения — у части
        # реальных товаров она оказалась битой, см. историю правки этой функции).
        if not isinstance(qty, (int, float)):
            header = str(name).strip()
            continue
        art = article(str(name))
        if not art:
            continue
        qty_by_article[art] = qty_by_article.get(art, 0) + int(qty)
        headers_by_article.setdefault(art, set()).add(header)

    stock_unambiguous = {a for a, hs in headers_by_article.items() if len(hs) == 1}
    safe_articles = price_unambiguous & stock_unambiguous
    stock = {a: qty_by_article[a] for a in safe_articles if a in qty_by_article}
    print(f"Остатки: источник {src.name}, {len(qty_by_article)} артикулов в файле, "
          f"{len(stock)} однозначных (безопасно сопоставлено).")
    return stock


def main() -> None:
    promo_meta = {}
    if PROMO_META_PATH.exists():
        promo_meta = json.loads(PROMO_META_PATH.read_text(encoding="utf-8"))
    else:
        print(f"ВНИМАНИЕ: сайдкар {PROMO_META_PATH} не найден - акционные товары останутся без oldPrice/discountPct.")

    categories = {}
    if CATEGORIES_PATH.exists():
        categories = json.loads(CATEGORIES_PATH.read_text(encoding="utf-8"))

    today = date.today()
    is_bootstrap = not FIRST_SEEN_PATH.exists()
    first_seen = {} if is_bootstrap else json.loads(FIRST_SEEN_PATH.read_text(encoding="utf-8"))
    # При самом первом запуске (файла ещё нет) весь текущий прайс - не "новинки", а то,
    # что уже давно продаётся, просто мы только сейчас начали это отслеживать. Ставим
    # дату заведомо за пределами окна NEW_WINDOW_DAYS, а не сегодняшнюю.
    bootstrap_date = (today - timedelta(days=NEW_WINDOW_DAYS + 1)).isoformat()
    if is_bootstrap:
        print(f"Первый запуск - журнал новинок {FIRST_SEEN_PATH} создаётся, текущий прайс новинками не считается.")

    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb.active

    items = []
    brand = None
    line = None
    idx = 0
    unknown_headers = []
    in_junk_section = False
    for row in ws.iter_rows(min_row=6):
        name_cell = row[1] if len(row) > 1 else None
        price_cell = row[2] if len(row) > 2 else None
        unit_cell = row[3] if len(row) > 3 else None
        category_cell = row[5] if len(row) > 5 else None
        if name_cell is None or name_cell.value is None:
            continue
        name = str(name_cell.value).strip()
        if not name:
            continue
        if is_excluded(name):
            continue
        price = price_cell.value if price_cell else None
        unit = unit_cell.value if unit_cell else None

        # Настоящий заголовок (бренд/линейка): нет ни цены, ни единицы измерения.
        if (price is None or not isinstance(price, (int, float))) and unit is None:
            if is_junk_section_header(name):
                in_junk_section = True
            else:
                in_junk_section = False
                level = header_level(name_cell)
                if level == "brand":
                    brand = name
                    line = None
                elif level == "line":
                    line = name
                else:
                    # Неопознанная заливка — считаем брендом (безопаснее, чем потерять товары внутри).
                    unknown_headers.append(name)
                    brand = name
                    line = None
            continue

        # Товар без цены (в т.ч. внутри мусорной секции выше) - пропускаем, не заголовок и не товар.
        if price is None or not isinstance(price, (int, float)):
            continue
        if in_junk_section:
            continue

        promo = is_red(name_cell) or (price_cell is not None and is_red(price_cell))
        idx += 1
        # round() убирает артефакты двоичного float (напр. 998.5799999999999 -> 998.58)
        price_rounded = round(float(price), 2)
        clean_name = re.sub(r"\s+", " ", name)
        item = {
            "id": idx,
            "brand": brand,
            "line": line,
            "name": clean_name,
            "price": int(price_rounded) if price_rounded.is_integer() else price_rounded,
            "promo": promo,
        }
        if promo:
            meta = promo_meta.get(clean_name)
            if meta:
                old_price = round(float(meta["old_price"]), 2)
                item["oldPrice"] = int(old_price) if old_price.is_integer() else old_price
                item["discountPct"] = meta["discount"]

        seen_date = first_seen.get(clean_name)
        if seen_date is None:
            seen_date = bootstrap_date if is_bootstrap else today.isoformat()
            first_seen[clean_name] = seen_date
        if (today - date.fromisoformat(seen_date)).days <= NEW_WINDOW_DAYS:
            item["isNew"] = True

        # Категория (см. CATEGORY_NAMES выше) - источник истины при каждом запуске это сама
        # колонка F текущего price-current.xlsx (пользователь мог поправить категорию вручную),
        # журнал categories обновляется отсюда же и просто переживает следующую пересборку
        # прайса. Если в колонке пусто - подстраховка журналом (на случай, если кто-то прогнал
        # build_price_current.py без add_category_column и разметка временно потерялась).
        cat_value = category_cell.value if category_cell else None
        cat_num = None
        if isinstance(cat_value, (int, float)) and int(cat_value) in CATEGORY_NAMES:
            cat_num = int(cat_value)
            categories[clean_name] = cat_num
        else:
            cat_num = categories.get(clean_name)
        if cat_num is not None and cat_num in CATEGORY_NAMES:
            item["category"] = CATEGORY_NAMES[cat_num]

        items.append(item)

    # Остатки на складе — отдельным проходом ПОСЛЕ сборки items (проверка на неоднозначность
    # артикула нужна против уже готового списка товаров, см. load_stock_levels).
    stock_levels = load_stock_levels(items)
    for it in items:
        qty = stock_levels.get(article(it["name"]))
        if qty is not None:
            it["stock"] = qty

    FIRST_SEEN_PATH.write_text(json.dumps(first_seen, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    CATEGORIES_PATH.write_text(json.dumps(categories, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")

    OUT_JSON.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    promo_count = sum(1 for i in items if i["promo"])
    with_old_price = sum(1 for i in items if i.get("oldPrice") is not None)
    new_count = sum(1 for i in items if i.get("isNew"))
    with_stock = sum(1 for i in items if i.get("stock") is not None)
    with_category = sum(1 for i in items if i.get("category"))
    without_category = [f'{i["brand"]} / {i["name"]}' for i in items if not i.get("category")]
    print(f"Сохранено {len(items)} товаров ({promo_count} акционных, из них {with_old_price} со старой ценой; "
          f"{new_count} новинок за последние {NEW_WINDOW_DAYS} дн.; {with_stock} с известным остатком на складе; "
          f"{with_category} с категорией) -> {OUT_JSON}")
    if without_category:
        print(f"Без категории (не попадут ни в одну категорию каталога, только в бренд): {without_category}")
    if promo_count and with_old_price < promo_count:
        print(f"ВНИМАНИЕ: {promo_count - with_old_price} акционных товаров не нашлись в сайдкаре по имени - проверить.")
    if unknown_headers:
        print(f"ВНИМАНИЕ: {len(unknown_headers)} заголовков с нераспознанной заливкой (взяты как бренд): {unknown_headers}")


if __name__ == "__main__":
    main()
