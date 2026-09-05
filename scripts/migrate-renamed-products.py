# -*- coding: utf-8 -*-
"""
Перенос журналов, когда товар в 1С сменил название ИЛИ артикул.

ПРАВИЛО ОТ ПОЛЬЗОВАТЕЛЯ (2026-09-05), на нём всё и держится:
    «товары могут менять названия или артикул, одновременно это не случается».
Значит опознать тот же товар можно надёжно:
    * совпал артикул, изменилось название  -> переименование;
    * совпало название, изменился артикул  -> смена артикула.
Если не совпало ни то, ни другое — это действительно новый товар.

Повод: в выгрузке от 5 сентября 2026 линейка Illumina сменила имя —
«Краска 10/05 Яркий блонд...» стала «Illumina 10/05 Яркий блонд...».

Почему без этого нельзя публиковать. Три журнала проекта привязаны к ИМЕНИ товара:
  * price-categories.json — категория товара (ручная разметка пользователя);
  * price-first-seen.json — когда товар впервые появился (из этого бейдж «Нов», 14 дней);
  * product-photos.json   — фото товара.
Без переноса переименованный товар теряет категорию (выпадает из режима «По категории»),
получает бейдж «Нов» — то есть сайт показывает покупателю неправду, — и теряет фото.

Запуск: .venv/Scripts/python.exe scripts/migrate-renamed-products.py [--apply]
Без --apply — только предпросмотр.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PRICE_XLSX = ROOT / 'public' / 'prices' / 'price-current.xlsx'
PRICE_ITEMS = ROOT / 'src' / 'data' / 'priceItems.json'
JOURNALS = [
    ROOT / 'src' / 'data' / 'price-categories.json',
    ROOT / 'src' / 'data' / 'price-first-seen.json',
    ROOT / 'src' / 'data' / 'product-photos.json',
]
# Палитры оттенков тоже держат ИМЯ товара из прайса — по нему каталог рисует кружок
# реального цвета рядом с товаром (swatchFor в PriceChecklist.astro). При переименовании
# связь рвётся молча: цвет просто перестаёт показываться. Поймано на Illumina 2026-09-05 —
# из 49 оттенков в каталоге осталось 3. В статьях палитра при этом продолжала работать,
# потому что resolveLiveShade ищет по АРТИКУЛУ, а не по имени, — расхождение и замаскировало
# проблему. Здесь чиним источник: переписываем имя в самом файле палитры.
SHADES = sorted((ROOT / 'src' / 'data').glob('*-shades.json'))


def norm(s) -> str:
    """Та же нормализация пробелов, что в xlsx-to-price-items.py."""
    return re.sub(r'\s+', ' ', str(s).strip())


def article_of(name: str) -> str:
    """Артикул — последний токен имени (та же логика, что в resolve-live-price.ts)."""
    toks = name.strip().split()
    return toks[-1] if toks else ''


def body_of(name: str) -> str:
    """Имя без артикула — то, что остаётся, когда меняется только артикул."""
    toks = name.strip().split()
    return ' '.join(toks[:-1]).lower() if len(toks) > 1 else name.strip().lower()


def read_new_price(known_brands: set):
    """Товары из свежего прайса вместе с брендом.

    Бренд определяем по СПИСКУ известных брендов, а не по заливке ячейки. Причина:
    add_category_column подсвечивает синим заголовки, под которыми есть неразмеченные
    товары, и эта подсветка затирает чёрную заливку бренда — определение по цвету
    начинает врать (поймано 2026-09-05: CONCEPT и KAPOUS переставали быть брендами).
    """
    ws = openpyxl.load_workbook(PRICE_XLSX).active
    items, brand = [], ''
    for r in range(6, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        price = ws.cell(row=r, column=3).value
        unit = ws.cell(row=r, column=4).value
        if not name:
            continue
        clean = norm(name)
        if price is None and not unit:                      # заголовок: бренд или линейка
            if clean.upper() in known_brands:
                brand = clean.upper()
            continue
        if price is not None:
            items.append({'name': clean, 'brand': brand})
    return items


def main() -> None:
    apply = '--apply' in sys.argv

    old_items = json.loads(PRICE_ITEMS.read_text(encoding='utf-8'))
    known_brands = {i['brand'].upper() for i in old_items}
    old_names = {i['name'] for i in old_items}

    new_items = read_new_price(known_brands)
    new_names = {i['name'] for i in new_items}
    print('Брендов распознано в новом прайсе: %d' % len({i['brand'] for i in new_items if i['brand']}))
    print('Товаров: было %d, стало %d' % (len(old_names), len(new_names)))

    gone = old_names - new_names
    added = [i for i in new_items if i['name'] not in old_names]

    # Индексы исчезнувших: по артикулу и по имени-без-артикула, в пределах бренда.
    by_article, by_body = {}, {}
    for i in old_items:
        if i['name'] in gone:
            b = i['brand'].upper()
            by_article.setdefault((b, article_of(i['name'])), []).append(i['name'])
            by_body.setdefault((b, body_of(i['name'])), []).append(i['name'])

    renames, ambiguous, truly_new = [], [], []
    for i in added:
        b = i['brand'].upper()
        same_article = by_article.get((b, article_of(i['name'])), [])
        same_body = by_body.get((b, body_of(i['name'])), [])
        cands = same_article or same_body
        why = 'артикул тот же' if same_article else 'название то же'
        if len(cands) == 1:
            renames.append((cands[0], i['name'], why))
        elif len(cands) > 1:
            ambiguous.append((i['name'], cands))
        else:
            truly_new.append(i['name'])

    print('\nОпознано как тот же товар: %d' % len(renames))
    for a, b, why in renames[:6]:
        print('   %s\n     -> %s   (%s)' % (a[:60], b[:60], why))
    if len(renames) > 6:
        print('   ... и ещё %d' % (len(renames) - 6))

    if ambiguous:
        print('\nНеоднозначные (несколько кандидатов — НЕ трогаем, решать человеку): %d' % len(ambiguous))
        for n, c in ambiguous[:6]:
            print('   %s <- %s' % (n[:52], [x[:32] for x in c]))

    print('\nДействительно новые товары: %d' % len(truly_new))
    for n in truly_new:
        print('   %s' % n[:72])

    truly_gone = sorted(gone - {a for a, _, _ in renames})
    print('\nДействительно ушли из прайса: %d' % len(truly_gone))
    for n in truly_gone[:6]:
        print('   %s' % n[:72])
    if len(truly_gone) > 6:
        print('   ... и ещё %d' % (len(truly_gone) - 6))

    if not renames:
        print('\nПереносить нечего.')
        return

    print()
    for path in JOURNALS:
        data = json.loads(path.read_text(encoding='utf-8'))
        moved = 0
        for old_name, new_name, _ in renames:
            if old_name in data and new_name not in data:
                data[new_name] = data.pop(old_name)
                moved += 1
        print('%-26s перенесено: %d' % (path.name, moved))
        if apply and moved:
            path.write_text(json.dumps(data, ensure_ascii=False, indent=1) + '\n', encoding='utf-8')

    rename_map = {a: b for a, b, _ in renames}
    for path in SHADES:
        shades = json.loads(path.read_text(encoding='utf-8'))
        moved = 0
        for s in shades:
            if s.get('name') in rename_map:
                s['name'] = rename_map[s['name']]
                moved += 1
        if moved:
            print('%-26s переименовано оттенков: %d' % (path.name, moved))
            if apply:
                path.write_text(json.dumps(shades, ensure_ascii=False, indent=1) + '\n', encoding='utf-8')

    print('\n' + ('Журналы и палитры обновлены.' if apply else 'Это предпросмотр. Для записи — с --apply'))


if __name__ == '__main__':
    main()
